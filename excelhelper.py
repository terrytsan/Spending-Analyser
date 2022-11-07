import openpyxl
import pandas
from openpyxl.styles.numbers import FORMAT_DATE_XLSX14
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


# Index is boolean, indicating if index should be included when writing df
def write_df_to_ws(data_frame: pandas.DataFrame, row_start: int, col_start: int, include_index: bool, ws: openpyxl.worksheet.worksheet.Worksheet):
    """
    Write a pandas data frame to worksheet
    
    :param data_frame: dataframe to write to worksheet
    :param row_start: row which data frame should start
    :param col_start: column which data frame should start
    :param include_index: True if data_frame's index should be included when writing df
    :param ws: worksheet in which dataframe should be inserted
    """
    if include_index:
        # Without this, there'll be an empty row between header and content when using dataframe_to_rows with index=True
        data_frame.reset_index(level=0, inplace=True)
    rows = dataframe_to_rows(data_frame, index=False)
    for r_idx, row in enumerate(rows, row_start):
        for c_idx, value in enumerate(row, col_start):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            if type(value) is pandas.Timestamp:
                cell.number_format = FORMAT_DATE_XLSX14
                ws.column_dimensions[cell.column_letter].auto_size = True
            
            column_name = data_frame.columns[c_idx-col_start]
            if 'cost' in column_name.lower():
                cell.number_format = '£#,##0.00'
    
    # Format as table
    table_start = get_column_letter(col_start) + str(row_start)
    table_height, table_width = data_frame.shape
    table_end = get_column_letter(col_start + table_width - 1) + str(row_start + table_height)
    
    tab = Table(displayName=f'Table{row_start}{col_start}{table_height}', ref=f'{table_start}:{table_end}')
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # Set columns to be auto_size
    for i in range(col_start, (col_start + table_width)):
        ws.column_dimensions[get_column_letter(i)].auto_size = True
