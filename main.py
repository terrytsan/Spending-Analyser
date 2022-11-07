# This is used for analysing and visualising spending data
from datetime import datetime
from functools import reduce

import openpyxl.worksheet.worksheet
import pandas
import pandas as pd
from typing import Tuple, List

from datetimerange import DateTimeRange
from openpyxl import load_workbook, Workbook
from openpyxl.chart import Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles.numbers import FORMAT_DATE_XLSX14
from openpyxl.utils import get_column_letter

from configparser import ConfigParser
from SpendAnalysis import SpendAnalysis
from excelhelper import write_df_to_ws

config = ConfigParser()
config.read('settings.ini')

filename = config.get('main', 'InputFileName')
output_filename = config.get('main', 'OutputFileName')
query_dates = [
    SpendAnalysis(datetime(2022, 10, 1), datetime(2022, 10, 31), "October"),
    SpendAnalysis(datetime(2022, 9, 1), datetime(2022, 9, 30), "September"),
    SpendAnalysis(datetime(2022, 8, 1), datetime(2022, 8, 31), "August"),
    SpendAnalysis(datetime(2022, 7, 1), datetime(2022, 7, 31), "July"),
    SpendAnalysis(datetime(2022, 6, 1), datetime(2022, 6, 30), "June"),
    SpendAnalysis(datetime(2022, 5, 1), datetime(2022, 5, 31), "May"),
    SpendAnalysis(datetime(2022, 4, 1), datetime(2022, 4, 30), "April"),
    SpendAnalysis(datetime(2022, 3, 1), datetime(2022, 3, 31), "March"),
]
output_sheet = "Analysis Output"
dataStartRow = 2
dataStartCol = 4


def get_table_as_dataframe(file_name: str, sheet_name: str, table_name: str) -> pandas.DataFrame:
    """
    Returns specified table as a pandas DataFrame.
    
    :param file_name: the path of the file containing the table
    :param sheet_name: name of sheet containing table
    :param table_name: name of table
    :return: table as pandas DataFrame
    """

    workbook = load_workbook(file_name, data_only=True)

    worksheet = workbook[sheet_name]

    table_ref = worksheet.tables[table_name].ref
    t_data = worksheet[table_ref]

    content = [[cell.value for cell in ent] for ent in t_data]

    header = content[0]
    rest = content[1:]

    return pd.DataFrame(rest, columns=header)


def is_intersecting(row: pandas.Series, query_start_date: datetime, query_end_date: datetime) -> bool:
    """
    Determines if row's 'Date Started' and 'Date Finished' columns are within
    bounds of query's start/end dates (inclusive).
    
    :param row: row under test
    :param query_start_date: lower bound (inclusive)
    :param query_end_date: upper bound (inclusive)
    :return: True if row is within bounds
    """
    if pd.isnull(row["Date Started"]):
        return False

    start_date = row["Date Started"].to_pydatetime()
    if start_date <= query_end_date and pd.isnull(row["Date Finished"]):
        return True

    query_range = DateTimeRange(query_start_date, query_end_date)
    test_range = DateTimeRange(row["Date Started"], row["Date Finished"])

    return query_range.is_intersection(test_range)


def analyse_data_in_dates(
        data: pandas.DataFrame,
        start_date: datetime,
        end_date: datetime
) -> Tuple[pandas.DataFrame, pandas.DataFrame, pandas.DataFrame]:
    """
    Considering only the data within the bounds set by start/end date (inclusive), generates and returns
    a Tuple containing the following data:
        - data_in_dates - data that have Date Started/Finished intersect with start_date and end_date
        - summary_by_category - data grouped by 'Category' field and summed.
        - unfinished_items - list of items where 'Date Finished' field is blank
    :param data: data to analyze
    :param start_date: date when analysis begins
    :param end_date: date when analysis ends
    :return: Tuple of the analysis results
    """

    # Filter only items with Date Started/Finished intersect with query dates
    filters = data.apply(is_intersecting, axis=1, args=(start_date, end_date))
    data_in_dates = data[filters].copy()

    data_in_dates["Days Used"] = (data_in_dates["Date Finished"] - data_in_dates["Date Started"]).dt.days + 1
    data_in_dates["Cost Per Day"] = data_in_dates["Cost"] / data_in_dates["Days Used"]

    # Remove unfinished items
    unfinished_items = data_in_dates[data_in_dates["Date Finished"].isnull()]
    data_in_dates = data_in_dates[data_in_dates["Date Finished"].notnull()]

    query_range = DateTimeRange(start_date, end_date)
    data_in_dates["Days Used in Query"] = data_in_dates.apply(lambda row: query_range.intersection(
        DateTimeRange(row["Date Started"], row["Date Finished"])).timedelta.days + 1, axis=1)

    data_in_dates["Cost in Query"] = data_in_dates["Cost Per Day"] * data_in_dates["Days Used in Query"]

    # Group and only keep 'Cost in Query' column
    summary_by_category = data_in_dates.groupby("Category").sum()[["Cost in Query"]]

    return data_in_dates, summary_by_category, unfinished_items


# Go through each query and create a summary table
def create_summary_table(sheet: openpyxl.worksheet.worksheet.Worksheet, query: List[SpendAnalysis]) -> None:
    """
    Creates a summary table in the specified sheet using the summary data within the query. summary_by_category field
    must be populated in each SpendAnalysis object. Run `analyse_data_in_dates`_ to populate.
    
    :param sheet: sheet to insert summary table
    :param query: list of SpendAnalysis objects with summary_by_category populated.
    """

    summary_df_array = []  # array containing all summary dfs
    for q in query:
        # Remove total row
        q.summary_by_category = q.summary_by_category.drop(
            q.summary_by_category[q.summary_by_category['Category'] == 'Total'].index)

        q.summary_by_category.rename(columns={"Cost in Query": q.alias + " Cost"}, inplace=True)
        summary_df_array.append(q.summary_by_category)

    complete_summary_df = reduce(lambda df_left, df_right: pd.merge(df_left, df_right, how='outer', on='Category'),
                                 summary_df_array)
    complete_summary_df.loc['Total'] = complete_summary_df.sum(numeric_only=True, axis=0)
    write_df_to_ws(complete_summary_df, 1, 1, False, sheet)


currentStartRow = dataStartRow
currentStartCol = dataStartCol

if __name__ == '__main__':
    df = get_table_as_dataframe(filename, "Spending", "Spending")

    wb = Workbook()
    ws = wb.create_sheet(output_sheet)

    for analysis in query_dates:
        result = analyse_data_in_dates(df, analysis.start_date, analysis.end_date)
        usedInQuery, categorySummary, unfinishedItems = result

        analysis.summary_by_category = categorySummary
        analysis.usedItems = usedInQuery
        analysis.unfinishedItems = unfinishedItems

        categorySummary.loc['Total'] = categorySummary.sum(numeric_only=True, axis=0)

        # Write back to excel

        # May have to delete everything in the sheet first
        # ws.delete_rows(1,ws.max_row+1);

        write_df_to_ws(categorySummary, currentStartRow, currentStartCol, True, ws)

        # Create pie chart for summary
        chart_location = get_column_letter(1) + str(currentStartRow + 10)
        cs_height, cs_width = categorySummary.shape
        pie = PieChart()
        pie_labels = Reference(ws, min_col=currentStartCol, min_row=currentStartRow + 1, max_col=currentStartCol,
                               max_row=currentStartRow + cs_height - 1)
        pie_data = Reference(ws, min_col=currentStartCol + cs_width - 1, min_row=currentStartRow,
                             max_col=currentStartCol + cs_width - 1,
                             max_row=currentStartRow + cs_height - 1)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(pie_labels)
        pie.title = f"Total by Category {analysis.start_date} - {analysis.end_date}"
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showVal = True
        ws.add_chart(pie, chart_location)

        currentStartCol = currentStartCol + categorySummary.shape[1] + 3
        write_df_to_ws(usedInQuery, currentStartRow, currentStartCol, False, ws)

        currentStartCol = currentStartCol + usedInQuery.shape[1] + 1
        write_df_to_ws(unfinishedItems, currentStartRow, currentStartCol, False, ws)

        # remove formatting
        # for row in ws.iter_rows(min_row=1, max_col=32, max_row=100):
        #     for cell in row:
        #         cell.style = 'Normal'

        ws.cell(row=currentStartRow, column=1).value = 'Start Date'
        ws.cell(row=currentStartRow + 1, column=1).value = 'End Date'
        ws.cell(row=currentStartRow, column=2).value = analysis.start_date
        ws.cell(row=currentStartRow, column=2).number_format = FORMAT_DATE_XLSX14
        ws.cell(row=currentStartRow + 1, column=2).value = analysis.end_date
        ws.cell(row=currentStartRow + 1, column=2).number_format = FORMAT_DATE_XLSX14
        ws.column_dimensions['B'].auto_size = True
        ws.column_dimensions['C'].auto_size = True

        # Start next date range 2 rows below previous content
        currentStartRow = currentStartRow + max(len(categorySummary), len(usedInQuery), len(unfinishedItems)) + 2
        currentStartCol = 4

    # Create the summary sheet
    summary_sheet = wb.create_sheet("Summary")
    create_summary_table(summary_sheet, query_dates)

    wb.remove(wb['Sheet'])
    wb.save(filename=output_filename)

    # Print additional info - not written to sheet
    unused = df[df['Date Started'].isnull()]
    totalUnused = unused['Cost'].sum()
    unfinished = df[df['Date Started'].notnull() & df['Date Finished'].isnull()]
    print("Total Unused: £", totalUnused)
    print("Total Unfinished: £", unfinished['Cost'].sum())
